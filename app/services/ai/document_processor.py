import os
import PyPDF2
import openpyxl
from typing import Optional

class DocumentProcessor:
    """Service for processing and extracting text from financial documents"""
    
    async def extract_text(self, file_path: str, file_extension: str) -> str:
        """Extract text content from various file formats"""
        
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        
        if file_extension == ".pdf":
            return await self._extract_pdf_text(file_path)
        elif file_extension == ".txt":
            return await self._extract_txt_text(file_path)
        elif file_extension == ".xlsx":
            return await self._extract_excel_text(file_path)
        elif file_extension == ".docx":
            return await self._extract_docx_text(file_path)
        else:
            raise ValueError(f"Unsupported file type: {file_extension}")
    
    async def _extract_pdf_text(self, file_path: str) -> str:
        """Extract text from PDF files"""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
                return text.strip()
        except Exception as e:
            raise Exception(f"Error extracting PDF text: {str(e)}")
    
    async def _extract_txt_text(self, file_path: str) -> str:
        """Extract text from TXT files"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read()
        except Exception as e:
            raise Exception(f"Error extracting TXT text: {str(e)}")
    
    async def _extract_excel_text(self, file_path: str) -> str:
        """Extract text from Excel files"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            text = ""
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"Sheet: {sheet_name}\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
                text += "\n"
            
            return text.strip()
        except Exception as e:
            raise Exception(f"Error extracting Excel text: {str(e)}")
    
    async def _extract_docx_text(self, file_path: str) -> str:
        """Extract text from DOCX files"""
        try:
            # For now, return a placeholder. In production, you'd use python-docx
            return f"DOCX content from {os.path.basename(file_path)} - Text extraction not implemented yet."
        except Exception as e:
            raise Exception(f"Error extracting DOCX text: {str(e)}")
    
    def extract_financial_metrics(self, text: str) -> dict:
        """Extract key financial metrics from text"""
        metrics = {
            "revenue": None,
            "profit": None,
            "earnings": None,
            "expenses": None,
            "cash_flow": None,
            "debt": None,
            "assets": None,
            "liabilities": None
        }
        
        # Simple keyword-based extraction
        lines = text.split('\n')
        for line in lines:
            line_lower = line.lower()
            
            # Revenue extraction
            if 'revenue' in line_lower and not metrics['revenue']:
                metrics['revenue'] = self._extract_number(line)
            
            # Profit extraction
            if any(word in line_lower for word in ['profit', 'net income', 'earnings']) and not metrics['profit']:
                metrics['profit'] = self._extract_number(line)
            
            # Expenses extraction
            if 'expense' in line_lower and not metrics['expenses']:
                metrics['expenses'] = self._extract_number(line)
        
        return metrics
    
    def _extract_number(self, text: str) -> Optional[float]:
        """Extract numeric values from text"""
        import re
        
        # Find numbers with currency symbols or commas
        number_pattern = r'[\$]?[\d,]+\.?\d*'
        matches = re.findall(number_pattern, text)
        
        if matches:
            # Clean up the first match
            number_str = matches[0].replace('$', '').replace(',', '')
            try:
                return float(number_str)
            except ValueError:
                return None
        
        return None 